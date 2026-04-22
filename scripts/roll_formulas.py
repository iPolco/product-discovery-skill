#!/usr/bin/env python3
"""Раскатывает формулы из колонки C (мес. 1) в D-N (мес. 2-12) на P&L и Cash Flow.

Шаблон financial-plan-template.xlsx содержит формулы только в колонке C. Колонки
D..N (мес. 2..12) пустые. Без раскатки финплан показывает выручку только за 1-й
месяц — для инвестора бесполезно.

Скрипт проходит по листам P&L и Cash Flow, находит строки с формулами в C и
раскатывает их в D..N со сдвигом относительных ссылок. Абсолютные ссылки ($X)
не трогаются.

Usage:
    python3 scripts/roll_formulas.py /home/claude/financial-plan.xlsx

После этого обязательно пересчитай:
    python3 /home/claude/scripts/recalc.py /home/claude/financial-plan.xlsx
"""
import sys
import re

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
except ImportError:
    print("❌ Нужна библиотека openpyxl. Установить: pip install openpyxl")
    sys.exit(1)


def shift_formula(formula: str, delta: int):
    """Сдвигает относительные ссылки на колонки на delta.

    Абсолютные ссылки ($C$5) не трогаются. Работает с двумя типами:
    - A1, AB5 — относительные, сдвигаются
    - $A1, $AB5 — абсолютные по колонке, не сдвигаются
    - A$1 — относительные по колонке, сдвигаются (строка фиксирована)
    """
    if not isinstance(formula, str) or not formula.startswith("="):
        return None

    def shift_ref(m):
        full = m.group(0)
        # Проверяем есть ли $ перед буквами колонки
        if full.startswith('$'):
            return full  # абсолютная по колонке — не трогаем
        col_letters = m.group(1)
        row_num = m.group(2)
        col_idx = column_index_from_string(col_letters)
        new_col_idx = col_idx + delta
        if new_col_idx < 1:
            return full
        return f"{get_column_letter(new_col_idx)}{row_num}"

    # Регекс для ссылок: опциональный $, буквы колонки, цифры строки
    return re.sub(r'\$?([A-Z]+)(\d+)', shift_ref, formula)


def roll_sheet(ws, start_col=4, end_col=15):
    """Раскатывает формулы из колонки C в колонки [start_col..end_col).

    По умолчанию D=4..N=14 включительно (end_col=15 — exclusive).

    Возвращает кортеж (rolled, skipped): сколько раскатано и сколько пропущено
    (потому что ячейка уже содержала значение).
    """
    rolled = 0
    skipped = 0
    for row in range(1, ws.max_row + 1):
        c_value = ws.cell(row=row, column=3).value
        if not isinstance(c_value, str) or not c_value.startswith("="):
            continue
        for col in range(start_col, end_col):
            if ws.cell(row=row, column=col).value is not None:
                skipped += 1
                continue
            shifted = shift_formula(c_value, col - 3)
            if shifted:
                ws.cell(row=row, column=col).value = shifted
                rolled += 1
    return rolled, skipped


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    path = sys.argv[1]

    try:
        wb = openpyxl.load_workbook(path)
    except FileNotFoundError:
        print(f"❌ Файл не найден: {path}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Не удалось открыть файл: {e}")
        sys.exit(1)

    total_rolled = 0
    for sheet_name in ["P&L", "Cash Flow"]:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  Лист «{sheet_name}» не найден — пропускаю")
            continue
        rolled, skipped = roll_sheet(wb[sheet_name])
        total_rolled += rolled
        print(f"✅ {sheet_name}: раскатано {rolled} формул"
              + (f" (пропущено {skipped} уже заполненных ячеек)" if skipped else ""))

    if total_rolled == 0:
        print()
        print("⚠️  Ничего не раскатано — возможно формулы уже на месте,")
        print("   или файл не соответствует ожидаемой структуре шаблона.")
        sys.exit(0)

    wb.save(path)
    print()
    print(f"✅ Сохранено: {path}")
    print()
    print("Следующий шаг — пересчитай значения через LibreOffice:")
    print(f"    python3 /home/claude/scripts/recalc.py {path}")
    print()
    print("Ожидаемый результат: \"status\": \"success\", \"total_errors\": 0")


if __name__ == "__main__":
    main()
