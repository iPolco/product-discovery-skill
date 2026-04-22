#!/usr/bin/env python3
"""Переставить лист 'Summary' первым в финплане.

Используется после заполнения финплана, чтобы инвестор/CEO видел executive-view
сразу при открытии файла, а не листал до 14-го листа.

Usage:
    python3 reorder_summary_first.py <path-to-financial-plan.xlsx>

Файл модифицируется на месте.
"""
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ Нужна библиотека openpyxl. Установить: pip install openpyxl")
    sys.exit(1)


def reorder_summary_first(xlsx_path: str, summary_sheet_name: str = "Summary") -> bool:
    """Переставить указанный лист первым. True если успешно, False если листа нет."""
    wb = load_workbook(xlsx_path)
    if summary_sheet_name not in wb.sheetnames:
        print(f"⚠️  Лист '{summary_sheet_name}' не найден в файле.")
        print(f"   Доступные листы: {wb.sheetnames}")
        return False

    # openpyxl хранит листы в wb._sheets (список) — перемещаем по индексу
    summary_sheet = wb[summary_sheet_name]
    current_index = wb._sheets.index(summary_sheet)
    if current_index == 0:
        print(f"✅ Лист '{summary_sheet_name}' уже первый, ничего делать не нужно.")
        return True

    # move_sheet с offset = -current_index переставит в начало
    wb.move_sheet(summary_sheet, offset=-current_index)
    wb.save(xlsx_path)
    print(f"✅ Лист '{summary_sheet_name}' перемещён с позиции {current_index + 1} на позицию 1.")
    return True


if __name__ == "__main__":
    if len(sys.argv) not in (2, 3):
        print(__doc__)
        sys.exit(1)

    path = Path(sys.argv[1])
    if not path.exists():
        print(f"❌ Файл не найден: {path}")
        sys.exit(1)

    sheet_name = sys.argv[2] if len(sys.argv) == 3 else "Summary"
    ok = reorder_summary_first(str(path), sheet_name)
    sys.exit(0 if ok else 1)
